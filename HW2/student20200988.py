#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']
scores = []

row_id = 1
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		scores.append(sum_v)
	row_id += 1

scores.sort(reverse=True)
count = len(scores)

row_id = 1
for row in ws:
	if row_id != 1:
		score = ws.cell(row = row_id, column = 7).value
		endIndex = scores.index(score)
		endIndex += scores.count(score)

		if endIndex <= count * 0.3:
			grade = 'A0'
		elif endIndex <= count * 0.7:
			grade = 'B0'
		else:
			grade = 'C0'

		ws.cell(row = row_id, column = 8).value = grade
	row_id += 1

wb.save("student.xlsx")

