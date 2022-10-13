import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

row_id = 1
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		ws.cell(row = row_id, column = 7).value = sum _v
	row_id += 1

wb.save("student.xlsx")
