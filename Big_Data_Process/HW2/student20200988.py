#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']
scores = []

row_id = 1
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		scores.append(sum_v)
	row_id += 1

scores.sort(reverse=True)
scores_length = len(scores)
grade_count = {'A0': 0, 'B0': 0, 'C0': 0}

row_id = 1
for row in ws:
	if row_id != 1:
		score = ws.cell(row = row_id, column = 7).value
		endIndex = scores.index(score)
		endIndex += scores.count(score)

		if endIndex <= scores_length * 0.3:
			grade = 'A0'
		elif endIndex <= scores_length * 0.7:
			grade = 'B0'
		else:
			grade = 'C0'

		grade_count[grade] += 1
		ws.cell(row = row_id, column = 8).value = grade

	row_id += 1

row_id = 1
for row in ws:
	if row_id != 1:
		score = ws.cell(row = row_id, column = 7).value
		endIndex = scores.index(score)
		endIndex += scores.count(score)

		if 0 <= endIndex and endIndex <= grade_count['A0'] * 0.5:
			ws.cell(row = row_id, column = 8).value = 'A+'
		elif grade_count['A0'] < endIndex and endIndex <= grade_count['A0'] + grade_count['B0'] * 0.5:
			ws.cell(row = row_id, column = 8).value = 'B+'
		elif grade_count['A0'] + grade_count['B0'] < endIndex and endIndex <= grade_count['A0'] + grade_count['B0'] + grade_count['C0'] * 0.5:
			ws.cell(row = row_id, column = 8).value = 'C+'
	row_id += 1

wb.save("student.xlsx")
#print(grade_count)
